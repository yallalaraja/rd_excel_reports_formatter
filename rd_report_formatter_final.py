#!/usr/bin/env python3
"""
RD Installment Report Formatter

Features:
- Accepts .xls or .xlsx input.
- Converts .xls to .xlsx using LibreOffice.
- Extracts original report details and RD account rows.
- Removes unwanted columns: E-Banking Ref No, Bank Name, Cheque Number,
  SB Account No, ASLAAS Number, Status, Last Created Date & Time.
- Adds S.No as the left-most table column.
- Preserves India Post logo if it exists in the converted workbook.
- Keeps RD Account Number leading zeroes.
- Applies formulas for Total Amount, Total No Of Records, and Total Deposit Amount.
- Applies bold text, center alignment, borders, spacing, and dynamic print setup.
- For small reports, fits to one A4 portrait page.
- For larger reports, fits width to one page and allows natural page height for readability.

Usage:
    python rd_report_formatter_final.py input.xls output.xlsx
    python rd_report_formatter_final.py input_folder output_folder

Requirements:
    pip install openpyxl
    LibreOffice must be installed for .xls conversion.
"""

from __future__ import annotations

import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.cell.cell import MergedCell


WANTED_HEADERS = [
    "RD Account Number",
    "Account Name",
    "RD Denomination",
    "RD Total Deposit Amount",
    "No of Installments",
    "Rebate",
    "Default fee",
]


def find_soffice() -> str:
    """Find LibreOffice/soffice on Windows/Linux/macOS even when PATH is not set."""
    candidates = [
        shutil.which("soffice"),
        shutil.which("libreoffice"),

        # Windows normal installer paths
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",

        # Common Linux/macOS paths
        "/usr/bin/soffice",
        "/usr/bin/libreoffice",
        "/snap/bin/libreoffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]

    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)

    raise RuntimeError(
        "LibreOffice is installed maybe, but soffice.exe was not found. "
        "Install LibreOffice from libreoffice.org or add this folder to PATH: "
        r"C:\Program Files\LibreOffice\program"
    )


def convert_xls_to_xlsx(input_path: Path) -> Path:
    """Convert .xls to .xlsx using LibreOffice. .xlsx files are returned unchanged."""
    if input_path.suffix.lower() != ".xls":
        return input_path

    office_cmd = find_soffice()
    temp_dir = Path(tempfile.mkdtemp(prefix="rd_report_convert_"))

    result = subprocess.run(
        [
            office_cmd,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(temp_dir),
            str(input_path),
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )

    if result.returncode != 0:
        raise RuntimeError(
            "LibreOffice conversion failed.\n"
            f"Command used: {office_cmd}\n"
            f"STDOUT: {result.stdout}\n"
            f"STDERR: {result.stderr}"
        )

    converted = temp_dir / f"{input_path.stem}.xlsx"
    if not converted.exists():
        # LibreOffice sometimes changes extension/name slightly, so search output folder.
        matches = list(temp_dir.glob("*.xlsx"))
        if matches:
            return matches[0]
        raise FileNotFoundError(
            f"Converted .xlsx file not found in {temp_dir}. "
            f"STDOUT: {result.stdout} STDERR: {result.stderr}"
        )
    return converted


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_text(value).lower())


def find_value_near_label(ws, label_keywords: List[str]) -> str:
    """Find a label cell and return the nearest non-empty value to its right."""
    label_keys = [normalize_key(x) for x in label_keywords]

    for row in ws.iter_rows():
        for cell in row:
            cell_key = normalize_key(cell.value)
            if not cell_key:
                continue
            if any(key in cell_key for key in label_keys):
                # Search right side in same row.
                for col in range(cell.column + 1, min(ws.max_column, cell.column + 12) + 1):
                    val = ws.cell(cell.row, col).value
                    if normalize_text(val):
                        return normalize_text(val)
    return ""


def find_header_row_and_cols(ws) -> tuple[int, Dict[str, int]]:
    """Find row containing RD table headers and return mapped columns."""
    for r in range(1, ws.max_row + 1):
        row_map: Dict[str, int] = {}
        row_text = []
        for c in range(1, ws.max_column + 1):
            val = normalize_text(ws.cell(r, c).value)
            if val:
                row_text.append(val.lower())
                for header in WANTED_HEADERS:
                    if normalize_key(header) == normalize_key(val):
                        row_map[header] = c
        joined = " ".join(row_text)
        if "rd account" in joined and "account name" in joined:
            # Some headers may have line breaks or spaces, so do fuzzy second pass.
            for c in range(1, ws.max_column + 1):
                val_key = normalize_key(ws.cell(r, c).value)
                for header in WANTED_HEADERS:
                    if normalize_key(header) == val_key:
                        row_map[header] = c
            if "RD Account Number" in row_map and "Account Name" in row_map:
                return r, row_map
    raise ValueError("Could not find RD table header row.")


def normalize_account_number(value: Any) -> str:
    """Return 12-digit account number text, preserving leading zeroes."""
    text = normalize_text(value).replace("'", "")
    text = re.sub(r"\.0$", "", text)
    digits = re.sub(r"\D", "", text)
    if digits:
        return digits.zfill(12)
    return text


def extract_data(ws) -> tuple[Dict[str, str], List[Dict[str, Any]]]:
    """Extract top report values and table data from the source sheet."""
    details = {
        "title": "RECURRING DEPOSIT INSTALLMENT REPORT",
        "agent": find_value_near_label(ws, ["Agent Id"]),
        "from_date": find_value_near_label(ws, ["From Date"]),
        "list_ref": find_value_near_label(ws, ["List Reference No", "List Ref No"]),
        "status": find_value_near_label(ws, ["Status"]),
        "cheque": find_value_near_label(ws, ["Cheque No"]),
        "type_report": find_value_near_label(ws, ["Type Of Report", "Type Report"]),
    }

    header_row, col_map = find_header_row_and_cols(ws)
    rd_col = col_map["RD Account Number"]
    name_col = col_map["Account Name"]

    data_rows: List[Dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        account = ws.cell(r, rd_col).value
        name = ws.cell(r, name_col).value
        if not normalize_text(account) and not normalize_text(name):
            continue

        # Stop when footer/total rows start.
        joined = " ".join(normalize_text(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1))
        if "total deposit amount" in joined or "total amount" in joined:
            break

        if normalize_text(account) and normalize_text(name):
            data_rows.append(
                {
                    "account": normalize_account_number(account),
                    "name": normalize_text(name),
                    "denom": ws.cell(r, col_map.get("RD Denomination", 0)).value if col_map.get("RD Denomination") else "",
                    "total": ws.cell(r, col_map.get("RD Total Deposit Amount", 0)).value if col_map.get("RD Total Deposit Amount") else "",
                    "installments": ws.cell(r, col_map.get("No of Installments", 0)).value if col_map.get("No of Installments") else "",
                    "rebate": ws.cell(r, col_map.get("Rebate", 0)).value if col_map.get("Rebate") else "",
                    "default_fee": ws.cell(r, col_map.get("Default fee", 0)).value if col_map.get("Default fee") else "",
                }
            )

    return details, data_rows


def clear_sheet_keep_images(ws):
    images = list(ws._images)
    for img in images:
        try:
            img.anchor = "A1"
        except Exception:
            pass

    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    max_rows = max(ws.max_row, 150)
    max_cols = max(ws.max_column, 40)
    for row in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.font = Font(name="Calibri", size=11)
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = "General"

    ws._images = images


def amount_to_formula(cell_ref: str) -> str:
    """Formula part to convert text like '5,000.00 Cr.' to number."""
    return f'IFERROR(VALUE(SUBSTITUTE(SUBSTITUTE({cell_ref}," Cr.",""),",","")),0)'



def apply_range_style(ws, cell_range: str, border=None, fill=None, font=None, alignment=None):
    """Apply style to every cell in a range, including merged-cell edge placeholders."""
    for row in ws[cell_range]:
        for cell in row:
            if border is not None:
                cell.border = border
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment


def apply_outline_border(ws, cell_range: str, side):
    """Draw a clean outside border around a rectangular range."""
    from openpyxl.utils.cell import range_boundaries

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            left = side if col == min_col else cell.border.left
            right = side if col == max_col else cell.border.right
            top = side if row == min_row else cell.border.top
            bottom = side if row == max_row else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def build_formatted_report(wb, ws, details: Dict[str, str], data_rows: List[Dict[str, Any]]):
    ws.title = "RDInstallmentReport"
    clear_sheet_keep_images(ws)

    row_count = len(data_rows)
    body_font_size = 14 if row_count <= 15 else 12 if row_count <= 35 else 11
    header_font_size = 14 if row_count <= 35 else 12

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)
    header_fill = PatternFill("solid", fgColor="C7C9FF")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    font_title = Font(name="Calibri", size=16, bold=True)
    font_header = Font(name="Calibri", size=header_font_size, bold=True)
    font_body = Font(name="Calibri", size=body_font_size, bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_nowrap = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Top fixed section.
    merge_ranges = [
        "A5:K5", "A6:K6",
        "A7:F7", "G7:K7", "A8:F8", "G8:K8", "A9:F9", "G9:K9",
        "A10:F10", "G10:K10", "A11:F11", "G11:K11", "A12:F12", "G12:K12",
        "A13:K13", "A14:F14", "G14:K14", "A15:F15", "G15:K15",
        "D17:F17", "G17:H17",
    ]
    for rng in merge_ranges:
        ws.merge_cells(rng)

    ws["A5"] = details.get("title") or "RECURRING DEPOSIT INSTALLMENT REPORT"
    ws["A6"] = "Search Criteria"
    top_rows = [
        ("A7", "Agent Id:", "G7", details.get("agent", "")),
        ("A8", "From Date:", "G8", details.get("from_date", "")),
        ("A9", "List Reference No:", "G9", details.get("list_ref", "")),
        ("A10", "Status:", "G10", details.get("status", "")),
        ("A11", "Cheque No.:", "G11", details.get("cheque", "")),
        ("A12", "Type Of Report:", "G12", details.get("type_report", "")),
    ]
    for label_cell, label_text, value_cell, value_text in top_rows:
        ws[label_cell] = label_text
        ws[value_cell] = value_text

    ws["A13"] = "Search Results"
    ws["A14"] = "Total Amount:"
    ws["A15"] = "Total No Of Records:"

    # Table headers.
    headers = {
        "A17": "S.No",
        "B17": "RD Account\nNumber",
        "C17": "Account Name",
        "D17": "RD Denomination",
        "G17": "RD Total Deposit\nAmount",
        "I17": "No of\nInstallments",
        "J17": "Rebate",
        "K17": "Default\nfee",
    }
    for cell, text in headers.items():
        ws[cell] = text

    start_row = 18
    for i, item in enumerate(data_rows, start=start_row):
        ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=6)
        ws.merge_cells(start_row=i, start_column=7, end_row=i, end_column=8)
        ws.cell(i, 1).value = f"=ROW()-{start_row - 1}"
        ws.cell(i, 2).value = item["account"]
        ws.cell(i, 2).number_format = "@"
        ws.cell(i, 3).value = item["name"]
        ws.cell(i, 4).value = item["denom"]
        ws.cell(i, 7).value = item["total"]
        ws.cell(i, 9).value = item["installments"]
        ws.cell(i, 10).value = item["rebate"] if item["rebate"] not in (None, "") else 0
        ws.cell(i, 11).value = item["default_fee"] if item["default_fee"] not in (None, "") else 0

    last_row = start_row + row_count - 1
    if row_count:
        parts = [amount_to_formula(f"G{r}") for r in range(start_row, last_row + 1)]
        sum_formula = f"=SUM({','.join(parts)})"
        count_formula = f"=COUNTA(B{start_row}:B{last_row})"
    else:
        sum_formula = "=0"
        count_formula = "=0"

    ws["G14"] = sum_formula
    ws["G15"] = count_formula

    # Gap between table and total block.
    total_header_row = last_row + 3 if row_count else start_row + 2
    total_value_row = total_header_row + 1

    ws.merge_cells(start_row=total_header_row, start_column=1, end_row=total_header_row, end_column=2)
    ws.merge_cells(start_row=total_value_row, start_column=1, end_row=total_value_row, end_column=2)
    ws.merge_cells(start_row=total_header_row, start_column=6, end_row=total_header_row, end_column=11)
    ws.merge_cells(start_row=total_value_row, start_column=6, end_row=total_value_row, end_column=11)
    ws.cell(total_header_row, 1).value = "List Ref No"
    ws.cell(total_value_row, 1).value = details.get("list_ref", "")
    ws.cell(total_header_row, 6).value = "Total Deposit Amount"
    ws.cell(total_value_row, 6).value = sum_formula

    used_max_row = total_value_row

    # General formatting.
    for row in ws.iter_rows(min_row=1, max_row=used_max_row, min_col=1, max_col=11):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.font = font_body
                cell.alignment = align_center
                cell.fill = white_fill

    # Clean borders for top section. Style merged ranges by range to avoid broken/partial
    # border lines in Excel Print Preview.
    top_style_ranges = [
        "A5:K5", "A6:K6",
        "A7:F7", "G7:K7", "A8:F8", "G8:K8", "A9:F9", "G9:K9",
        "A10:F10", "G10:K10", "A11:F11", "G11:K11", "A12:F12", "G12:K12",
        "A13:K13", "A14:F14", "G14:K14", "A15:F15", "G15:K15",
    ]
    for rng in top_style_ranges:
        apply_range_style(ws, rng, border=border_all, fill=white_fill, font=font_body, alignment=align_center)
        apply_outline_border(ws, rng, thin)

    ws["A5"].font = font_title
    for rng in ["A6:K6", "A13:K13"]:
        apply_range_style(ws, rng, fill=header_fill, font=font_header, alignment=align_center)
        apply_outline_border(ws, rng, thin)

    # Table header borders/fill.
    for row in ws.iter_rows(min_row=17, max_row=17, min_col=1, max_col=11):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.fill = header_fill
                cell.border = border_medium
                cell.font = font_header
                cell.alignment = align_center

    # Table data formatting.
    for r in range(start_row, last_row + 1):
        for c in range(1, 12):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.border = border_all
                cell.font = font_body
                cell.fill = white_fill
                cell.alignment = align_nowrap if c in [2, 3, 4, 7] else align_center

    # Total block formatting. Only style the two real blocks, not the blank gap,
    # so Print Preview stays clean and simple.
    for rng in [
        f"A{total_header_row}:B{total_header_row}",
        f"A{total_value_row}:B{total_value_row}",
        f"F{total_header_row}:K{total_header_row}",
        f"F{total_value_row}:K{total_value_row}",
    ]:
        apply_range_style(ws, rng, border=border_all, fill=white_fill, font=font_body, alignment=align_center)
        apply_outline_border(ws, rng, thin)

    for rng in [f"A{total_header_row}:B{total_header_row}", f"F{total_header_row}:K{total_header_row}"]:
        apply_range_style(ws, rng, fill=header_fill, font=font_header, alignment=align_center)
        apply_outline_border(ws, rng, thin)

    ws["G14"].number_format = "#,##0.00"
    ws["G15"].number_format = "0"
    ws.cell(total_value_row, 6).number_format = "#,##0.00"

    # Widths: start from A, no empty A-E columns.
    widths = {
        "A": 7, "B": 18, "C": 39, "D": 12, "E": 8, "F": 8,
        "G": 14, "H": 8, "I": 13, "J": 9, "K": 9,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
        ws.column_dimensions[col].hidden = False
    for c in range(12, 50):
        col = get_column_letter(c)
        ws.column_dimensions[col].hidden = True
        ws.column_dimensions[col].width = 0.1

    # Row heights.
    for r in range(1, used_max_row + 1):
        ws.row_dimensions[r].height = 22
    for r, h in {1: 55, 2: 55, 3: 30, 4: 10, 5: 30, 6: 26, 16: 12, 17: 52}.items():
        ws.row_dimensions[r].height = h
    data_height = 35 if body_font_size >= 14 else 27
    for r in range(start_row, last_row + 1):
        ws.row_dimensions[r].height = data_height
    ws.row_dimensions[total_header_row - 1].height = 18
    ws.row_dimensions[total_header_row].height = 32
    ws.row_dimensions[total_value_row].height = 32

    # View and print setup.
    ws.sheet_view.view = "normal"
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1 if row_count <= 20 else 0
    ws.page_margins = PageMargins(left=0.10, right=0.10, top=0.15, bottom=0.15, header=0.03, footer=0.03)
    ws.print_options.horizontalCentered = True
    ws.print_area = f"A1:K{used_max_row}"
    ws.row_breaks.brk = []
    ws.col_breaks.brk = []
    ws.freeze_panes = None

    # Reposition logo if image exists.
    for img in ws._images:
        try:
            img.anchor = "A1"
        except Exception:
            pass

    return {"account_rows": row_count, "print_area": ws.print_area}


def format_report(input_file: str | Path, output_file: str | Path) -> Dict[str, Any]:
    input_path = Path(input_file)
    output_path = Path(output_file)
    source_xlsx = convert_xls_to_xlsx(input_path)

    wb = load_workbook(source_xlsx)
    ws = wb.active
    details, data_rows = extract_data(ws)
    info = build_formatted_report(wb, ws, details, data_rows)
    wb.save(output_path)
    return info


def batch_format(input_folder: str | Path, output_folder: str | Path) -> List[Path]:
    input_folder = Path(input_folder)
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)

    outputs: List[Path] = []
    for file in sorted(input_folder.iterdir()):
        if file.suffix.lower() not in [".xls", ".xlsx"]:
            continue
        out = output_folder / f"{file.stem}_formatted.xlsx"
        format_report(file, out)
        outputs.append(out)
    return outputs


def main():
    if len(sys.argv) != 3:
        print("Usage:")
        print("  python rd_report_formatter_final.py input.xls output.xlsx")
        print("  python rd_report_formatter_final.py input_folder output_folder")
        raise SystemExit(1)

    src = Path(sys.argv[1])
    dst = Path(sys.argv[2])

    if src.is_dir():
        outputs = batch_format(src, dst)
        print(f"Formatted {len(outputs)} file(s). Output folder: {dst}")
    else:
        info = format_report(src, dst)
        print(f"Created: {dst}")
        print(f"Detected account rows: {info['account_rows']}")
        print(f"Print area: {info['print_area']}")


if __name__ == "__main__":
    main()
